#!/usr/bin/env python
# coding: utf-8

# # EasyCellCounting.py
# 
# This program iterates through flourescent cell images and counts cells on the left and right side of selected regions of interest. While this program was initially created for Tomato fluorescing neurons in the spinal cord and brain, EasyCellCounting.py can be used for any images that require efficient, easy counting of flourescent cells.  
# <a id="0"></a> <br>
# ## Table of Contents  
# 1. [Packages used by EasyCellCounting.py](#1)     
# 1. [Useful Functions: the Rotation and InsertionSort Functions](#2) 
# 1. [Creating two lists to store big TIFF images and the smaller, resized images](#3) 
#     1. [Activating an emtpy Excel file called "Input.xlsx" to use later and instantiating needed arrays](#4) 
#     1. [Array Instantiation and Definitions](#5) 
# 1. [Iterating through small images for user input](#6)     
# 1. [Iterating and Counting through Large TIFF Files Autonomously](#7)     
# 1. [Add counts to a blank Excel file and color code the data](#8)
#     1. [Color Code Definition](#9) 

# <a id="1"></a> <br>
# ## Packages used by EasyCellCounting.py:
# 
# These python packages are required for EasyCellCounting.py to work:

# In[1]:


# Used to link folders and files to code
from os import listdir
from os.path import isfile, join


# Used to create arrays and perform mathematical operations
import numpy as np
import pandas as pd

#Packages used to edit and view images:
import cv2
import PIL
from PIL import Image

#Packages used to open and edit Excel Files
from openpyxl import load_workbook
from openpyxl.styles import Font
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color


# <a id="2"></a> <br>
# ## Useful Functions: the Rotation and InsertionSort Functions
# 
# These functions are used frequently throughout the rest of the code.

# In[2]:


# This function rotates an image given an angle and an input image
def rotate_image(image, angle):
    Center = tuple(np.array(image.shape[1::-1]) / 2)
    Matrix = cv2.getRotationMatrix2D(Center, angle, 1.0)
    # Utilizes openCV rotation Matrix function to rotate images 
    rotated_image = cv2.warpAffine(image, Matrix, image.shape[1::-1], flags=cv2.INTER_LINEAR)
    return rotated_image
        

# The insertionSort function performs insertion sort to organize files in sequential (ascending) order

def insertionSort(arr1, arr2):
     
    if (n := len(arr1)) <= 1:
        return
    
    for i in range(1, n):
         
        key1 = arr1[i]
        key2 = arr2[i]
        m = i-1
        while m >=0 and key1 < arr1[m] :
                arr1[m+1] = arr1[m]
                arr2[m+1] = arr2[m]
                m -= 1
        arr1[m+1] = key1
        arr2[m+1] = key2
        # Sorts through each file one-by-one by looking at the NAMES of each file.
        # The "<" operation forces the program to look for the first different character in the names of the file.
        # One can use this fact to name all files the same format except for certain numbers (to demarcate order)
        # For example, when comparing "Zymo6 3730.4 1-12_s1.tif" and "Zymo6 3730.4 1-12_s2.tif", the program will place "Zymo6 3730.4 1-12_s1.tif" first in the onlyfiles array.
        # This is due to naming of "s1" (in this case, "s1" means sample #1).
        


# <a id="4"></a> <br>
# ## Creating two lists to store big TIFF images and the smaller, resized images
# 
# The program also checks to see if each large TIFF file has its corresponding smaller image 
# 
# 

# <a id="4"></a> <br>
# ## Activating an emtpy Excel file called "Input.xlsx" to use later and instantiating needed arrays
# 
# EasyCellCounting.py will insert all data into this blank excel file:

# In[3]:


#Creating list "r_images" for large TIFF files
r_mypath= "/Users/amav/Documents/My Programs :)/Pics"
r_onlyfiles = [ f for f in listdir(r_mypath) if isfile(join(r_mypath,f)) ]
r_files = [ f for f in listdir(r_mypath) if isfile(join(r_mypath,f)) ]

for i in range(0, len(r_files)):
    r_files[i] = r_onlyfiles[i]
insertionSort(r_files, r_onlyfiles)
r_images = np.empty(len(r_onlyfiles), dtype=object)

for i in range(0, len(r_onlyfiles)):
    print(r_onlyfiles[i])


#Creating list "images" for smaller files
mypath= "/Users/amav/Documents/My Programs :)/Pics (lower)"
onlyfiles = [ f for f in listdir(mypath) if isfile(join(mypath,f)) ]
files = [ f for f in listdir(mypath) if isfile(join(mypath,f)) ]

for i in range(0, len(files)):
    files[i] = onlyfiles[i]

insertionSort(files, onlyfiles)
images = np.empty(len(onlyfiles), dtype=object)


for i in range(0, len(files)):
    if r_onlyfiles[i] not in files:
        print(files[i])

#Making sure the big TIFF list is the same size as the small images list because each reduced, small image needs to have its corresponding big TIFF image. 
assert len(r_images) == len(images)


# In[4]:


# Activate blank Excel file: 
test_file = "input.xlsx"
obj = openpyxl.load_workbook(test_file)
sheet = obj.active


# <a id="5"></a> <br>
# ### These arrays will store final images and counts to display to user and insert into activated excel file respectively. 
# 
# The "finished" array will store selected region of interests with counted cells higlighted in green (and blue for identified clusters). The "final_count" array will store the names of each image that is counted.

# In[5]:


finished = []
final_count = []


# ### These arrays will store data used in the program
# 
# For example, these arrays store the position of selected boxes, user inputed / default brighness levels, and user inputed minimum area / default minimum area of neuron (these inputs are optional to the user).

# In[6]:


# Minimum area to classify as a neuron / cell array: 
min_area_list = []
cluster_max = []

# These arrays will store the left and right counts for each image respectively:
left_count = []
right_count = []

# These arrays will help store the final locations of the selected regions of interest
xpos1 = []
xpos2 = []
ypos1 = []
ypos2 = []

# This array will store the brightness index for each image (defaulting to 190)
bright = []


# <a id="6"></a> <br>
# # Iterating through small images for user input
# 
# ### Here are the user input options below:
# 
# 
# Press "T" to select box and press space to confirm box
# 
# Press "U" as many times as needed to move backwards (and redo images)
# 
# Press "W, "E", or "O" to rotate counterclockwise, clockwise, or a full 180 degrees respectively
# 
# 
# Press "M" to break loop
# 
# Press "Q" to check for the green (you'll need to enter the desired brightness value to check)
# 
# Press "V" to use 200 for the brightness threshold
# 
# #### The user must iterate through every small image. (EX: 100 images takes approximately 3 minutes)

# In[ ]:


# Starting point
n = 0
p = 0

while n < len(onlyfiles):

    images[n] = cv2.imread(join(mypath,onlyfiles[n]))
    r_images[n] = cv2.imread(join(r_mypath,r_onlyfiles[n]))
     
    piet = images[n]
    
    while True:
    
        cv2.imshow(str(n) + " / " + str(len(onlyfiles)) + " " + str(onlyfiles[n]), piet)


        key = cv2.waitKey(1) & 0xFF
        
        #User input conditionals:
            
        if key == ord('w'):
            piet = rotate_image(piet, 7)
            r_images[n] = rotate_image(r_images[n], 7)
            cv2.destroyAllWindows()
            
        if key == ord('e'):
            piet = rotate_image(piet, -7)
            r_images[n] = rotate_image(r_images[n], -7)
            cv2.destroyAllWindows()
            
        if key == ord('o'):
            piet = rotate_image(piet, 180)
            r_images[n] = rotate_image(r_images[n], 180)
            cv2.destroyAllWindows()
        
        if key == ord('t'):
            break
        if key == ord('m'):
            break
        if key == ord('u'):
            break
        if key == ord('q'):
            break
        if key == ord('v'):
            break
        if key == ord('p'):
            break
        if key == ord('i'):
            break
    
    if key == ord('u'):
        # If the "u" key is pressed, the program erases all information on the previous image and allows the user to redo the box selection
        
        xpos1.pop()
        xpos2.pop()
        ypos1.pop()
        ypos2.pop()
        bright.pop()
        min_area_list.pop()
        cluster_max.pop()
        
        # Changing the iteration index here:
        n = n - 1
        continue
        
    if key == ord('q'):
        
        #Ask user for the input value (if nothing is inputed then default brightness at 160)
        
        bright_temp = int(input("Enter brightness value: "))
        min_area = 40
        cv2.destroyAllWindows()
        #Perform the identification of neurons algorithm
        
        # Selecting Region of Interest
        r = cv2.selectROI("select the area", piet)

        # Cropping the image to selected box
        cropped_image = piet[int(r[1]):int(r[1]+r[3]), 
                              int(r[0]):int(r[0]+r[2])]

        # Finding middle of selected box
        middle = r[2] / 2

        x1, y1 = int(r[0]), int(r[1])
        x2, y2 = int(r[0]+r[2]), int(r[1]+r[3])

        ratio = len(r_images[n]) / len(piet)
        ratio2 = r_images[n].shape[1] / piet.shape[1]

        # Reproportioning small boxes to the larger TIFF images
        nx1, ny1 = int(x1 * ratio), int(y1 * ratio2)
        nx2, ny2 = int(x2 * ratio), int(y2 * ratio2)
        
        #Now to process the image:
        
        piat = r_images[n]
        cropped_image = piat[ny1:ny2, nx1:nx2]

        # Identifying the middle to later sort counts into left / right arrays
        middle = (nx2 - nx1) / 2

        # Converting the cropped image to RGB format to allow for PIL manipulations
        color_coverted = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2RGB)
        image = PIL.Image.fromarray(color_coverted)
        pix = image.load()

        width, height = image.size

        # Contrast filter: interating through each pixel of the cropped_image. 
        # Due to the relatively small area of the cropped_image, this does not take too much time. 
        for y in range(height):
            for x in range(width):
                r = pix[x,y][0]
                rng = range(bright_temp, 255, 1)
                if r in rng:
                    pix[x, y] = (255, 0, 0)
                else:
                    pix[x, y] = (0, 0, 0)

        cvimg = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)

        img = cv2.cvtColor(cvimg, cv2.COLOR_BGR2GRAY)

        # Finding countours of potential cells to count: 
        cnts = cv2.findContours(img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = cnts[0] if len(cnts) == 2 else cnts[1]


        # Initializing left / right arrays for the specific image as well as the max_area constant.
        # The max area constant can be approximated by the user after a few trail images are counted. 
        max_area = 900

        area_list = []

        # Counting with OpenCV Countour function

        for c in cnts:
            (x,y),radius = cv2.minEnclosingCircle(c)
            area = cv2.contourArea(c)
            area_list.append(area)
            if area > min_area and area < max_area:
                # Drawing green countour over counted cells: 
                cv2.drawContours(cropped_image, [c], -1, (0, 250, 0), 2)


        # CLUSTER ALGORITHM:
        if len(area_list) > 3: # Requires at least 3 normal cells to be present in the image
            Biggest_Cell_area = np.percentile(area_list, 99)

            for c1 in cnts:
                area = cv2.contourArea(c1)
                if area > max_area and area < 10000:
                    # Drawing a blue contour around identified clusters of cells: 
                    cv2.drawContours(cropped_image, [c1], -1, (250, 0, 0), 3)
        
        #Show the processed image
        cluster_max.append(10000)
        cv2.imshow("Check", cropped_image)
        cv2.waitKey()

        cv2.destroyAllWindows()
        
        #Go back to the same image to select the box again (same algorithm as the u key one)
        continue
            
            
    cv2.destroyAllWindows()
    
    # Selecting Region of Interest
    r = cv2.selectROI("select the area", piet)
    
    # Cropping the image to selected box
    cropped_image = piet[int(r[1]):int(r[1]+r[3]), 
                          int(r[0]):int(r[0]+r[2])]
    
    # Finding middle of selected box
    middle = r[2] / 2
    
    x1, y1 = int(r[0]), int(r[1])
    x2, y2 = int(r[0]+r[2]), int(r[1]+r[3])
    
    ratio = len(r_images[n]) / len(piet)
    ratio2 = r_images[n].shape[1] / piet.shape[1]
    
    # Reproportioning small boxes to the larger TIFF images
    nx1, ny1 = int(x1 * ratio), int(y1 * ratio2)
    nx2, ny2 = int(x2 * ratio), int(y2 * ratio2)
    
    # Updating values to array lists
    
    xpos1.append(nx1)
    xpos2.append(nx2)
    ypos1.append(ny1)
    ypos2.append(ny2)
    
    if key == ord('t'):  
        bright.append(160)
        min_area_list.append(40)
        cluster_max.append(10000)
    elif key == ord('p'):
        bright.append(250)
        min_area_list.append(40)
        cluster_max.append(10000)
    elif key == ord('v'):
        bright.append(200)
        min_area_list.append(40)
        cluster_max.append(10000)
    elif key == ord('i'):
        temp = int(input("Brightness value:"))
        bright.append(temp)
        min_area_list.append(40)
        cluster_max.append(1000000)
    else:
        bright.append(160)
        min_area_list.append(40)
        cluster_max.append(10000)
        
    if key == ord('m'):
            break
            
    # Give user warning if no box is selected for an image (User will need to redo this image by pressing "U" key)
    if y2 == 0:     
        print("Caution Error at " + str(n))
    
    n += 1
 


# <a id="7"></a> <br>
# # Iterating and Counting through Large TIFF Files Autonomously
# 
# This For Loop iterates through all the larger TIFF files and applies filters to count the cells and identify clusters of cells.
# 
# ## 1. Contrast Filter to remove all background noise and filter all bright cells within a brightness range
# The program then converts the filtered image back into an OpenCV file to perform later operations.
# 
# ## 2. Cluster Algorithm:
# 
# For clusters of cells, the program identifies any areas of bright flourescent color with an area of greater than 900 pixels as a cluster (900 is a constant). The program procceeds to get the area of the biggest cell to try and fit as many of that big cell into the area of the cluster. In short, the program divides the area of the cluster by the area of the single biggest cell and rounds up. That number is then added to the corresponding left or right counts (and reported to the user)
# 

# In[ ]:


for n in range(0, len(xpos1)):

    piat = r_images[n]
    cropped_image = piat[ypos1[n]:ypos2[n], xpos1[n]:xpos2[n]]
    
    # Identifying the middle to later sort counts into left / right arrays
    middle = (xpos2[n] - xpos1[n]) / 2
    
    # Converting the cropped image to RGB format to allow for PIL manipulations
    color_coverted = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2RGB)
    image = PIL.Image.fromarray(color_coverted)
    pix = image.load()
    
    width, height = image.size
    
    # Contrast filter: interating through each pixel of the cropped_image. 
    # Due to the relatively small area of the cropped_image, this does not take too much time. 
    for y in range(height):
        for x in range(width):
            r = pix[x,y][0]
            rng = range(bright[n], 255, 1)
            if r in rng:
                pix[x, y] = (255, 0, 0)
            else:
                pix[x, y] = (0, 0, 0)

    cvimg = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)

    img = cv2.cvtColor(cvimg, cv2.COLOR_BGR2GRAY)
    
    # Finding countours of potential cells to count: 
    cnts = cv2.findContours(img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cnts = cnts[0] if len(cnts) == 2 else cnts[1]
    
    
    # Initializing left / right arrays for the specific image as well as the max_area constant.
    # The max area constant can be approximated by the user after a few trail images are counted. 
    max_area = 900
    left_white_dots = []
    right_white_dots = []

    
    area_list = []
    
    # Counting with OpenCV Countour function
    
    for c in cnts:
        area = cv2.contourArea(c)
        if area > min_area_list[n] and area < max_area:
            # Drawing green countour over counted cells: 
            cv2.drawContours(cropped_image, [c], -1, (0, 250, 0), 2)
            # Extracting useful information about the location of each countour in coordinates:
            (x,y),radius = cv2.minEnclosingCircle(c)
            # Sorting which cells are on the left and right side of the middle: 
            if x < middle:
                left_white_dots.append(c)
            elif x > middle:
                right_white_dots.append(c)
            area_list.append(area)
            
    
    # CLUSTER ALGORITHM:

    
    if len(area_list) > 3: # Requires at least 3 normal cells to be present in the image
        Biggest_Cell_area = np.percentile(area_list, 99)

        for c1 in cnts:
            area = cv2.contourArea(c1)
            if area > max_area and area < cluster_max[n]:
                # Drawing a blue contour around identified clusters of cells: 
                cv2.drawContours(cropped_image, [c1], -1, (250, 0, 0), 2)
                (x,y),radius = cv2.minEnclosingCircle(c1)
                # Check stores the minimum number of cells within the identified clusters:
                check = int((area / Biggest_Cell_area) + 1)
                
                # Adding the cluster counts to each respective side and reporting this to the user:
                if x < middle:
                    for i in range(0, check):
                        left_white_dots.append(c1)
                    print("Added to left: ", check)
                elif x > middle:
                    for i in range(0, check):
                        right_white_dots.append(c1)
                    print("Added to right: ", check)
                
    print(onlyfiles[n])       
    print("Left count - Right count: ",len(left_white_dots), len(right_white_dots))
    
    final_count.append(onlyfiles[n])
    left_count.append(len(left_white_dots))
    right_count.append(len(right_white_dots))
    
    

    finished.append(cropped_image)
    cv2.destroyAllWindows()


# ### Display regions of interests with counted cells only if there are less than 10 images analyzed
#  If there are more than 10 images analyzed, there may be too many images to display for the computer, resulting in overloading.

# In[ ]:


if len(finished) < 10:
    for n in range(0, len(finished)):
        cv2.imshow(onlyfiles[i], finished[n])
        cv2.waitKey()

print(len(xpos1))


# <a id="8"></a> <br>
# # Add counts to a blank Excel file and color code the data
# 
# This program sorts through the blank Excel input and enters all the left counts in the left counts table and the right counts in the right count table. The program also color codes the data based on the number and region of spinal cord.
# 
# The specific color code of the data depends on the naming nomenclature. In our case we use 3556.1 s_1. "3556.1" represents the mouse number and "s_1" represents the first spinal cord from that mouse number. All spinal cords from the same mouse have the same highlighted color in the Excel file. 

# In[ ]:


for n in range(0, len(onlyfiles)):
    final_count[n] = onlyfiles[n]


# This is simply the name of the first image:
y = onlyfiles[0]

# The index value [5:13] represent the string "3730.4 "from the initial name "Zymo6 3730.4 1-12_s11.tif"
# "3730.4 " represents the tag of the animal from which the section was collected. 
# Hence, these index values below will change according to the user's naming choice
name = y[6:13] #NAMING

# Row counter and Coloumn counter
sheet.cell(row = 1, column = 1).value = "Left Counts (NEED TO CHECK BLUE DYE)"
sheet.cell(row = 36, column = 1).value = "Right Counts (NEED TO CHECK BLUE DYE)"

for i in range(1, 30):
    sheet.cell(row = 2, column = i + 1).value = i
    sheet.cell(row = 37, column = i + 1).value = i

r = 3
c = 1
full = '\t'.join(onlyfiles)
sheet.cell(row = r, column = c).value = name
sheet.cell(row = r + 35, column = c).value = name

print(onlyfiles)
print(left_count)
print(right_count)
print(name in final_count[0])


for n in range(0, len(onlyfiles)):
    # If the section is from the same animal as before, continue on the same row (each animal has its own row)
    if name in onlyfiles[n]:
        c = c + 1
        # Add the left count to the top table
        sheet.cell(row = r, column = c).value = left_count[n]
        # Add the right count to the bottom table
        sheet.cell(row = r + 35, column = c).value = right_count[n]
        
        # Conditional statements for color coding (via highlighting):
        # If sections are from different regions of the brain / spinal cord, the user can enter A1 - E4 codes in the file name to allow for each color coding
        if "A1" in final_count[n]:
            color = 27
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "A2" in final_count[n]:
            color = 44
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "A3" in final_count[n]:
            color = 49
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "A4" in final_count[n]:
            color = 4
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "B1" in final_count[n]:
            color = 5
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "B2" in final_count[n]:
            color = 50
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "B3" in final_count[n]:
            color = 57
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "B4" in final_count[n]:
            color = 19
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "C1" in final_count[n]:
            color = 45
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "C2" in final_count[n]:
            color = 29
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "C3" in final_count[n]:
            color = 22
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "C4" in final_count[n]:
            color = 23
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
    
    # Else if there is a new animal tag, move on to the next row.
    
    else:
        c = 1
        r = r + 1
        name = final_count[n][6:13] #NAMING
        sheet.cell(row = r, column = c).value = name
        sheet.cell(row = r + 35, column = c).value = name
        c = c + 1
        sheet.cell(row = r, column = c).value = left_count[n]
        sheet.cell(row = r + 35, column = c).value = right_count[n]
        
        
        
        
        # Same color coding technique here: 
        if "A1" in final_count[n]:
            color = 27
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "A2" in final_count[n]:
            color = 44
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "A3" in final_count[n]:
            color = 49
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "A4" in final_count[n]:
            color = 4
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "B1" in final_count[n]:
            color = 5
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "B2" in final_count[n]:
            color = 50
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "B3" in final_count[n]:
            color = 57
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "B4" in final_count[n]:
            color = 19
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "C1" in final_count[n]:
            color = 45
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "C2" in final_count[n]:
            color = 29
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "C3" in final_count[n]:
            color = 22
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp
        elif "C4" in final_count[n]:
            color = 23
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
            sheet.cell(row = r, column = c).fill = temp
            sheet.cell(row = r + 35, column = c).fill = temp


# <a id="9"></a> <br>
# ## Display the Color Code on Row 60 of the Excel document

# In[ ]:


sheet.cell(row = 90, column = 1).value = "Color code: "
sheet.cell(row = 90, column = 2).value = "A1"
color = 27
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 2).fill = temp

sheet.cell(row = 90, column = 3).value = "A2"
color = 44
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 3).fill = temp

sheet.cell(row = 90, column = 4).value = "A3"
color = 49
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 4).fill = temp

sheet.cell(row = 90, column = 5).value = "A4"
color = 4
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 5).fill = temp

sheet.cell(row = 90, column = 6).value = "B1"
color = 5
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 6).fill = temp

sheet.cell(row = 90, column = 7).value = "B2"
color = 50
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 7).fill = temp

sheet.cell(row = 90, column = 8).value = "B3"
color = 57
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 8).fill = temp

sheet.cell(row = 90, column = 9).value = "B4"
color = 19
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 9).fill = temp

sheet.cell(row = 90, column = 10).value = "C1"
color = 45
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 10).fill = temp

sheet.cell(row = 90, column = 11).value = "C2"
color = 29
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 11).fill = temp

sheet.cell(row = 90, column = 12).value = "C3"
color = 22
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 12).fill = temp

sheet.cell(row = 90, column = 13).value = "C4"
color = 23
temp = PatternFill(patternType='solid', fgColor=Color(indexed=color))
sheet.cell(row = 90, column = 13).fill = temp

            
# Save the results in a new file called "cell_counts.xlsx"
obj.save(filename="cell_counts.xlsx")
print("Completed")


# In[ ]:


yo = "CFA 30.14 01-12_s1.tif"
yo2 = yo[8:15]
print(yo2)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




